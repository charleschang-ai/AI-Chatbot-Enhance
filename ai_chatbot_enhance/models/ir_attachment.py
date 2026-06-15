import base64
import contextlib
import csv
import io
import logging
import re

from odoo import models, api
from ..models.models import AI_SUPPORTED_IMG_TYPES
from odoo.tools.pdf import OdooPdfFileReader, OdooPdfFileWriter, to_pdf_stream, PdfReadError
from odoo.tools.image import ImageProcess

_logger = logging.getLogger(__name__)

from odoo.tools.lru import LRU
import zipfile
import xml.dom.minidom
import warnings
from lxml import etree
import importlib.util


if not (importlib.util.find_spec('pdfminer') and importlib.util.find_spec('pdfminer.high_level')):
    _logger.warning("Attachment indexation of PDF documents is unavailable because the 'pdfminer.six' Python library cannot be found on the system. "
                    "You may install it from https://pypi.org/project/pdfminer.six/ (e.g. `pip3 install pdfminer.six`)")


FTYPES = ['docx', 'pptx', 'xlsx', 'opendoc', 'pdf']
index_content_cache = LRU(1)


def textToString(element):
    buff = u""
    for node in element.childNodes:
        if node.nodeType == xml.dom.Node.TEXT_NODE:
            buff += node.nodeValue
        elif node.nodeType == xml.dom.Node.ELEMENT_NODE:
            buff += textToString(node)
    return buff


def _csv_escape(value):
    if value is None:
        return ''
    value = str(value)
    if ',' in value or '"' in value or '\n' in value or '\r' in value:
        return '"' + value.replace('"', '""') + '"'
    return value


def _clean_text_content(buf):
    """Clean PDF content: remove NULs, normalize whitespace and line breaks."""
    if not buf:
        return buf
    # Remove NULs, normalize CRLF/CR to LF, replace tabs with spaces
    buf = buf.translate({
        ord('\x00'): None,
        ord('\r'): None,
        ord('\t'): ord(' '),
    })

    # Collapse runs of whitespace while preserving at most a single blank line
    def _compact_whitespace(match):
        chunk = match.group(0)
        newline_count = chunk.count('\n')
        if newline_count == 0:
            return ' '
        return '\n\n' if newline_count > 1 else '\n'

    buf = re.sub(r'\s{2,}', _compact_whitespace, buf)
    return buf.strip()


class IrAttachment(models.Model):
    _inherit = 'ir.attachment'

    AI_MAX_PDF_PAGES = 5
    TABULAR_FILE_TYPES = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # xlsx
        'application/vnd.ms-excel',  # xls
        'application/vnd.oasis.opendocument.spreadsheet',  # ods
        'text/csv',  # csv
    ]

    def _index_docx(self, bin_data):
        '''Index Microsoft .docx documents'''
        buf = u""
        f = io.BytesIO(bin_data)
        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                content = xml.dom.minidom.parseString(zf.read("word/document.xml"))
                for val in ["w:p", "w:h", "text:list"]:
                    for element in content.getElementsByTagName(val):
                        buf += textToString(element) + "\n"
            except Exception:
                pass
        return buf

    def _index_pptx(self, bin_data):
        '''Index Microsoft .pptx documents'''

        buf = u""
        f = io.BytesIO(bin_data)
        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                zf_filelist = [x for x in zf.namelist() if x.startswith('ppt/slides/slide')]
                for i in range(1, len(zf_filelist) + 1):
                    content = xml.dom.minidom.parseString(zf.read('ppt/slides/slide%s.xml' % i))
                    for val in ["a:t"]:
                        for element in content.getElementsByTagName(val):
                            buf += textToString(element) + "\n"
            except Exception:
                pass
        return buf

    def _index_xlsx(self, bin_data):
        '''Index Microsoft .xlsx documents'''

        try:
            from openpyxl import load_workbook  # noqa: PLC0415
            logging.getLogger("openpyxl").setLevel(logging.CRITICAL)
        except ImportError:
            _logger.info('openpyxl is not installed.')
            return ""

        f = io.BytesIO(bin_data)
        all_sheets = []
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(f, data_only=True, read_only=True)
                for sheet in workbook.worksheets:
                    sheet_name = sheet.title
                    sheet_name_escaped = _csv_escape(sheet_name)
                    sheet_rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        row_cells = [sheet_name_escaped] + [
                            _csv_escape(str(cell) if cell is not None else '') for cell in row
                        ]
                        sheet_rows.append(','.join(row_cells))
                    sheet_data = '\n'.join(sheet_rows)
                    if sheet_data:
                        all_sheets.append(sheet_data)
        except Exception:  # noqa: BLE001
            pass

        all_sheets_str = '\n\n'.join(all_sheets)
        return _clean_text_content(all_sheets_str)

    @api.model
    def _index(self, bin_data, mimetype, checksum=None):
        if checksum:
            cached_content = index_content_cache.get(checksum)
            if cached_content:
                return cached_content
        res = False
        for ftype in FTYPES:
            buf = getattr(self, '_index_%s' % ftype)(bin_data)
            if buf:
                res = buf.replace('\x00', '')
                break

        res = res or super(IrAttachment, self)._index(bin_data, mimetype, checksum=checksum)
        if checksum:
            index_content_cache[checksum] = res
        return res

    def _index_opendoc(self, bin_data):
        '''Index OpenDocument documents (.odt, .ods...)'''

        f = io.BytesIO(bin_data)
        buf = []
        MAX_COLUMN_REPEAT = 100
        MAX_ROW_REPEAT = 50
        main_namespaces = {
            'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
            'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
            'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
            'manifest': 'urn:oasis:names:tc:opendocument:xmlns:manifest:1.0'
        }

        def extract_row(row):
            cells = []
            for cell in row.xpath('.//table:table-cell | .//table:covered-table-cell', namespaces=main_namespaces):
                repeat = cell.get(f'{{{main_namespaces["table"]}}}number-columns-repeated')
                repeat_count = min(int(repeat), MAX_COLUMN_REPEAT) if repeat and repeat.isdigit() else 1
                text_parts = cell.xpath('.//text:p//text()', namespaces=main_namespaces)
                cell_text = ' '.join(t.strip() for t in text_parts if t.strip())
                cells.extend([cell_text] * repeat_count)
            return cells

        def extract_spreadsheet(content):
            sheets_csv = []
            tables = content.xpath('.//table:table', namespaces=main_namespaces)
            for table in tables:
                table_rows = []
                table_name = table.get(f'{{{main_namespaces["table"]}}}name')
                if not table_name:
                    table_name = f"Sheet{len(sheets_csv) + 1}"
                table_name_escaped = _csv_escape(table_name)
                for row in table.xpath('.//table:table-row', namespaces=main_namespaces):
                    row_repeat = row.get(f'{{{main_namespaces["table"]}}}number-rows-repeated')
                    row_repeat_count = min(int(row_repeat), MAX_ROW_REPEAT) if row_repeat and row_repeat.isdigit() else 1

                    cells = extract_row(row)
                    if not any(cells):
                        continue

                    while cells and not cells[-1]:
                        cells.pop()

                    row_str = ','.join([table_name_escaped] + list(map(_csv_escape, cells)))
                    if row_str.replace(',', '').strip():
                        table_rows.extend([row_str] * row_repeat_count)

                if table_rows:
                    sheets_csv.append('\n'.join(table_rows))

            return sheets_csv

        def extract_text(content):
            lines = []
            for element in content.xpath('.//text:p | .//text:h | .//text:list-item', namespaces=main_namespaces):
                text = ''.join(element.xpath('.//text()', namespaces=main_namespaces)).strip()
                if text:
                    lines.append(text)
            return lines

        if zipfile.is_zipfile(f):
            try:
                zf = zipfile.ZipFile(f)
                content = etree.fromstring(zf.read('content.xml'))
                mime_type = zf.read('mimetype').decode('utf-8').strip()
                if mime_type and 'spreadsheet' in mime_type:
                    buf.extend(extract_spreadsheet(content))
                else:
                    buf.extend(extract_text(content))
            except Exception:
                pass

        buf_str = '\n\n'.join(buf)
        return _clean_text_content(buf_str)

    def _index_pdf(self, bin_data):
        '''Index PDF documents'''
        if not bin_data.startswith(b'%PDF-'):
            return ""
        try:
            if not importlib.util.find_spec('pdfminer.high_level'):
                return ""
            from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter  # noqa: PLC0415
            from pdfminer.converter import TextConverter  # noqa: PLC0415
            from pdfminer.layout import LAParams  # noqa: PLC0415
            from pdfminer.pdfpage import PDFPage  # noqa: PLC0415
            logging.getLogger("pdfminer").setLevel(logging.CRITICAL)
        except ImportError:
            # warned already during init of module
            return ""
        f = io.BytesIO(bin_data)
        try:
            resource_manager = PDFResourceManager()
            # Setting boxes_flow triggers the _group_textboxes function,
            # used to group textboxes by distance, which helps sort them
            # better. In our case, we don't need to sort them this way,
            # so we can disable the feature to reduce the memory footprint
            # of the library and avoid memory issues on most PDF files.
            laparams = LAParams(detect_vertical=True, boxes_flow=None)

            with io.StringIO() as content, TextConverter(
                resource_manager,
                content,
                laparams=laparams
            ) as device:
                interpreter = PDFPageInterpreter(resource_manager, device)
                for page in PDFPage.get_pages(f):
                    interpreter.process_page(page)

                buf = content.getvalue()
            return _clean_text_content(buf)
        except Exception:  # noqa: BLE001
            return ""

    def _get_ai_attachment_content(self):
        """
        Get the indexing-processed content of the attachment
        """
        self.ensure_one()
        attachment_content = ''
        if self.mimetype in self.TABULAR_FILE_TYPES:
            sheets = self.index_content.split('\n\n')
            for sheet in sheets:
                if sheet:
                    rows_list = self._process_csv_text(sheet)
                    if rows_list:
                        result = '\n'.join(str(row) for row in rows_list)
                        attachment_content += result + '\n'
        else:
            attachment_content = self.index_content

        if not attachment_content:
            return None

        # Check for reasonable content length and word variety
        if len(attachment_content.strip()) <= 10:
            return None

        unique_words = {w.lower() for w in attachment_content.split()}
        if len(unique_words) < 2:
            return None

        return attachment_content

    def _setup_ai_attachment_chunks(self, embedding_model, content=None):
        self.ensure_one()
        if self.mimetype in self.TABULAR_FILE_TYPES:
            chunks = content.split('\n')
        else:
            chunks = self._chunk_text(content)

        vals_list = []
        for chunk in chunks:
            if self.name:
                chunk = f"Attachment Name: {self.name}\n{chunk}"
            vals_list.append({
                'attachment_id': self.id,
                'content': chunk,
                'embedding_model': embedding_model,
            })

        self.env['ai.embedding'].create(vals_list)

    @staticmethod
    def _process_csv_text(csv_text):
        """
        Process CSV text into a list of dictionaries with headers as keys.
        :return: List of row dictionaries or None if invalid
        :rtype: list[dict] or None
        """

        lines = csv_text.strip().split('\n')
        if not lines:
            return None

        # Detect delimiter and header
        sample = '\n'.join(lines[:min(10, len(lines))])

        delimiter = ','
        has_header = False
        try:
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            has_header = sniffer.has_header(sample)
        except csv.Error:
            pass

        # Generate headers from first row or create generic ones
        if has_header:
            first_row = next(csv.reader(io.StringIO(lines[0]), delimiter=delimiter))
            headers = [h.strip() if h else f"Column_{i}" for i, h in enumerate(first_row)]
        else:
            first_row = next(csv.reader(io.StringIO(lines[0]), delimiter=delimiter))
            headers = [f"Column_{i}" for i in range(len(first_row))]

        # Parse CSV with safety nets for ragged rows
        reader = csv.DictReader(
            io.StringIO(csv_text),
            delimiter=delimiter,
            fieldnames=headers,
            restkey='_extra_fields',  # Extra columns to be added as extra fields key
            restval=None  # Missing columns will be added as None
        )

        if has_header:
            next(reader, None)

        rows = list(reader)
        return rows

    @staticmethod
    def _clean_text(text):
        """
        Clean up the text content while preserving meaningful structure.

        :param str text: Raw text content to clean
        :returns: Cleaned text content
        :rtype: str
        """
        # Remove NUL characters that can cause PostgreSQL insertion errors
        text = text.replace('\x00', '')
        text = text.replace('\r\n', '\n')

        # Split into lines and process
        lines = text.split('\n')
        result = []
        current_paragraph = []

        for line in lines:
            stripped = line.strip()

            if not stripped:
                if current_paragraph:
                    result.append(' '.join(current_paragraph))
                    current_paragraph = []
                continue

            if stripped.endswith(':') or (len(stripped) < 120 and stripped.endswith('.')):
                if current_paragraph:
                    result.append(' '.join(current_paragraph))
                    current_paragraph = []
                result.append(stripped)
                continue

            current_paragraph.append(stripped)

        if current_paragraph:
            result.append(' '.join(current_paragraph))

        # Join with single newlines between paragraphs
        text = '\n'.join(result)

        # Clean up extra whitespace
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'\s+([.,!?;:])', r'\1', text)
        return text.strip()

    @staticmethod
    def _chunk_text(text, chunk_size=1500, margin=200, min_chunk_size=1000, max_chunk_size=5000):
        """
        Split text into chunks based on character count with a hard maximum limit.
        The hard max limit is 5000 characters so that the chunks have enough context for the LLM to understand in case of large chunks.

        :param str text: The input text to chunk.
        :param int chunk_size: Target chunk size in characters.
        :param int margin: Allow flexibility in chunk sizes within chunk_size ± margin.
        :param int min_chunk_size: Minimum size a chunk should have before finalizing.
        :param int max_chunk_size: Hard maximum size limit that cannot be exceeded.
        :return: List of text chunks
        :rtype: list[str]
        """
        cleaned_text = IrAttachment._clean_text(text)
        chunks = []
        paragraphs = cleaned_text.split('\n')

        current_chunk = []
        current_length = 0

        def _add_chunk_enforcing_max_size(chunk_content):
            """Add a chunk, splitting it if it exceeds max_chunk_size."""
            if len(chunk_content) <= max_chunk_size:
                chunks.append(chunk_content)
            else:
                # Force split oversized chunks by words
                words = chunk_content.split()
                temp_chunk = []
                temp_length = 0
                for word in words:
                    word_length = len(word) + 1
                    if temp_length + word_length > max_chunk_size:
                        if temp_chunk:
                            chunks.append(" ".join(temp_chunk))
                        temp_chunk = [word]
                        temp_length = len(word)
                    else:
                        temp_chunk.append(word)
                        temp_length += word_length
                if temp_chunk:
                    chunks.append(" ".join(temp_chunk))

        for para in paragraphs:
            para = para.strip()
            if not para:
                continue  # Skip empty lines

            para_length = len(para)

            # If current chunk is too small, try to merge with the next one
            if current_chunk and (current_length + para_length + 1 <= chunk_size + margin):
                current_chunk.append(para)
                current_length += para_length + 1
                continue

            # If the current chunk is large enough, store it
            if current_length >= chunk_size - margin:
                _add_chunk_enforcing_max_size(" ".join(current_chunk))
                current_chunk = [para]  # Start a new chunk
                current_length = para_length
            else:
                # If chunk is too small but para itself is too big, split it
                if para_length > chunk_size:
                    # First, store current chunk if it exists
                    if current_chunk:
                        _add_chunk_enforcing_max_size(" ".join(current_chunk))
                        current_chunk = []
                        current_length = 0
                    # Split the large paragraph by sentences
                    sentences = re.split(r'(?<=[.!?])\s+', para)
                    temp_chunk = []
                    temp_length = 0

                    for sentence in sentences:
                        sent_length = len(sentence)

                        if temp_length + sent_length + 1 > chunk_size:
                            if temp_chunk:
                                _add_chunk_enforcing_max_size(" ".join(temp_chunk))
                            temp_chunk = [sentence]
                            temp_length = sent_length
                        else:
                            temp_chunk.append(sentence)
                            temp_length += sent_length + 1

                    if temp_chunk:
                        _add_chunk_enforcing_max_size(" ".join(temp_chunk))
                else:
                    current_chunk.append(para)
                    current_length += para_length + 1

        # Handle the last chunk
        if current_chunk:
            last_chunk_content = " ".join(current_chunk)
            if chunks and len(last_chunk_content) < min_chunk_size:
                # Try to merge with the previous chunk, but respect max_chunk_size
                if len(chunks[-1]) + len(last_chunk_content) + 1 <= max_chunk_size:
                    chunks[-1] += " " + last_chunk_content
                else:
                    # Can't merge, add as separate chunk
                    _add_chunk_enforcing_max_size(last_chunk_content)
            else:
                _add_chunk_enforcing_max_size(last_chunk_content)

        return chunks

    def _ai_read(self, fnames, files_dict):
        """When attachments are inserted in a prompt, one send the files (or indexed contents) to
        the LLMs.
        """
        if fnames:
            return super()._ai_read(fnames, files_dict)
        vals = []
        for attachment in self:
            if attachment.checksum in files_dict:
                vals.append({'id': attachment.id, 'file': files_dict[attachment.checksum]['file_ref']})
                continue
            file_ref = f'<file_#{len(files_dict) + 1}>'
            extension = attachment.mimetype.split('/')[-1]
            if extension == 'pdf' and not attachment.url:
                # Extract the X first / last pages of the PDFs
                reader = None
                with contextlib.suppress(PdfReadError):
                    reader = OdooPdfFileReader(to_pdf_stream(attachment), strict=False)
                if not reader or reader.numPages <= self.AI_MAX_PDF_PAGES:
                    b64_datas = attachment.datas.decode()
                else:
                    writer = OdooPdfFileWriter()
                    start_pages = self.AI_MAX_PDF_PAGES // 2
                    end_pages = self.AI_MAX_PDF_PAGES - start_pages
                    for p in (*range(start_pages), *range(reader.numPages - end_pages, reader.numPages)):
                        writer.addPage(reader.getPage(p))
                    out_buff = io.BytesIO()
                    writer.write(out_buff)
                    b64_datas = base64.b64encode(out_buff.getvalue()).decode()

                files_dict[attachment.checksum] = {
                    'mimetype': 'application/pdf',
                    'value': b64_datas,
                    'file_ref': file_ref,
                }
            elif extension in AI_SUPPORTED_IMG_TYPES and not attachment.url:
                raw_data = attachment.raw

                try:
                    image_process = ImageProcess(raw_data)
                    size = image_process.image.size
                    if max(size) > 1024:
                        raw_data = image_process \
                            .crop_resize(min(size[0], 1024), min(size[1], 1024), 0, 0) \
                            .image_quality(output_format='PNG')
                except Exception as e:  # noqa: BLE001
                    _logger.error("Image resize failed %s", e)

                files_dict[attachment.checksum] = {
                    'mimetype': attachment.mimetype,
                    'value': base64.b64encode(raw_data).decode(),
                    'file_ref': file_ref,
                }
            else:
                if not attachment.index_content or attachment.index_content == "application":
                    continue
                files_dict[attachment.checksum] = {
                    'mimetype': 'text/plain',
                    'value': attachment.index_content,
                    'file_ref': file_ref
                }
            vals.append({'id': attachment.id, 'file': file_ref})
        return vals, files_dict
